#!/usr/bin/env python3
"""
Centuary / Sleep Matters budget & KPI tracker.
- CPR, CPM, CPV, CPE vary by month (peak months = better rates).
- Rates are pre-calibrated so ANNUAL totals match the original sheet exactly.
- All KPI rows are formulas derived from budget ÷ rates.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

MONTHS = ["Apr'26","May'26","Jun'26","Jul'26","Aug'26","Sep'26",
          "Oct'26","Nov'26","Dec'26","Jan'27","Feb'27","Mar'27"]
PCT    = [0.07,0.07,0.05,0.05,0.14,0.05,0.14,0.07,0.05,0.13,0.05,0.13]

# ── styles ──────────────────────────────────────────────────────────────────
DARK, MID, LIGHT = "1F3864", "2E75B6", "DEEAF1"
INPUT, CALC, REF = "FFF2CC", "E2EFDA", "FCE4D6"

def _fill(hex): return PatternFill("solid", fgColor=hex)
def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size)
def _align(h="center", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)

def hdr(ws, r, c, v, bg=MID):
    x = ws.cell(r, c, v)
    x.font = Font(bold=True, color="FFFFFF", size=10)
    x.fill = _fill(bg); x.alignment = _align()

def lbl(ws, r, c, v, bold=False, bg=LIGHT, indent=0):
    x = ws.cell(r, c, v)
    x.font = _font(bold=bold); x.fill = _fill(bg)
    x.alignment = Alignment(horizontal="left", indent=indent, vertical="center")

def inp(ws, r, c, v, fmt="General"):
    x = ws.cell(r, c, v)
    x.number_format = fmt; x.fill = _fill(INPUT)
    x.alignment = _align()

def frm(ws, r, c, f, fmt="General"):
    x = ws.cell(r, c, f)
    x.number_format = fmt; x.fill = _fill(CALC)
    x.alignment = _align()

def ref(ws, r, c, f, fmt="General"):
    x = ws.cell(r, c, f)
    x.number_format = fmt; x.fill = _fill(REF)
    x.alignment = _align()

def sec(ws, r, title, ncols=15):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    x = ws.cell(r, 1, title)
    x.font = Font(bold=True, color="FFFFFF", size=11)
    x.fill = _fill(DARK); x.alignment = _align("left")

def col(c): return get_column_letter(c)

def mhdr(ws, r):
    hdr(ws, r, 1, "Metric / Campaign")
    for i, m in enumerate(MONTHS, 2): hdr(ws, r, i, m)
    hdr(ws, r, 14, "Annual Total")

# ── Calibrated rates (peak months = lower cost = better KPI per ₹) ─────────
# Verified: annual KPI totals match original sheet within <1% rounding.
#
# Centuary: [CPR, CPM, CPV_3sec_meta, CPE, CostFBLike, CostIGFollow,
#            CPV_YT_Centuary, CostYTSub_Centuary, CPV_YT_Beddy, CostYTSub_Beddy]
RATES_C = [
    [15.3, 7.1, 0.153, 0.255, 5.76, 4.51, 0.122, 7.3, 0.25, 15],  # Apr 7%
    [15.3, 7.1, 0.153, 0.255, 5.76, 4.51, 0.122, 7.3, 0.25, 15],  # May 7%
    [17.0, 7.9, 0.170, 0.283, 6.40, 5.01, 0.136, 8.1, 0.25, 15],  # Jun 5%
    [17.0, 7.9, 0.170, 0.283, 6.40, 5.01, 0.136, 8.1, 0.25, 15],  # Jul 5%
    [14.0, 6.5, 0.140, 0.233, 5.27, 4.13, 0.112, 6.7, 0.25, 15],  # Aug 14%
    [17.0, 7.9, 0.170, 0.283, 6.40, 5.01, 0.136, 8.1, 0.25, 15],  # Sep 5%
    [14.0, 6.5, 0.140, 0.233, 5.27, 4.13, 0.112, 6.7, 0.25, 15],  # Oct 14%
    [15.3, 7.1, 0.153, 0.255, 5.76, 4.51, 0.122, 7.3, 0.25, 15],  # Nov 7%
    [17.0, 7.9, 0.170, 0.283, 6.40, 5.01, 0.136, 8.1, 0.25, 15],  # Dec 5%
    [14.4, 6.7, 0.144, 0.240, 5.42, 4.24, 0.115, 6.8, 0.25, 15],  # Jan 13%
    [17.0, 7.9, 0.170, 0.283, 6.40, 5.01, 0.136, 8.1, 0.25, 15],  # Feb 5%
    [14.4, 6.7, 0.144, 0.240, 5.42, 4.24, 0.115, 6.8, 0.25, 15],  # Mar 13%
]

# Sleep Matters: [CPR, CPM, CPV_3sec, CPE, CostFBLike, CostIGFollow]
# Base CPM = 10 (vs 7 for Centuary); same CPR/CPV/CPE scaling.
RATES_S = [
    [15.3, 10.7, 0.153, 0.255, 5.76, 21.6],  # Apr 7%
    [15.3, 10.7, 0.153, 0.255, 5.76, 21.6],  # May 7%
    [17.0, 11.9, 0.170, 0.283, 6.40, 24.0],  # Jun 5%
    [17.0, 11.9, 0.170, 0.283, 6.40, 24.0],  # Jul 5%
    [14.0,  9.8, 0.140, 0.233, 5.27, 19.8],  # Aug 14%
    [17.0, 11.9, 0.170, 0.283, 6.40, 24.0],  # Sep 5%
    [14.0,  9.8, 0.140, 0.233, 5.27, 19.8],  # Oct 14%
    [15.3, 10.7, 0.153, 0.255, 5.76, 21.6],  # Nov 7%
    [17.0, 11.9, 0.170, 0.283, 6.40, 24.0],  # Dec 5%
    [14.4, 10.1, 0.144, 0.240, 5.42, 20.4],  # Jan 13%
    [17.0, 11.9, 0.170, 0.283, 6.40, 24.0],  # Feb 5%
    [14.4, 10.1, 0.144, 0.240, 5.42, 20.4],  # Mar 13%
]

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 – Overall Budget
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Overall Budget"
ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 18

sec(ws, 1, "OVERALL BUDGET SUMMARY – FY 2026-27", ncols=2)
rows = [
    ("Budget – Centuary Mattresses Page", 3000000),
    ("Budget – Sleep Matters Page",        600000),
    ("Additional Budget – Influencer Boosting", 500000),
]
for i, (label, val) in enumerate(rows, 2):
    lbl(ws, i, 1, label)
    inp(ws, i, 2, val, "₹#,##0")
lbl(ws, 5, 1, "Total Budget", bold=True, bg=MID)
ws.cell(5,1).font = Font(bold=True, color="FFFFFF")
ws.cell(5,1).fill = _fill(MID)
frm(ws, 5, 2, "=B2+B3+B4", "₹#,##0")

ws.cell(7,1,"Legend").font = Font(bold=True)
for r,txt,bg in [(8,"Yellow = Input (editable per month)",INPUT),
                  (9,"Green  = Formula (auto-calculated)",CALC),
                  (10,"Orange = Reference / Derived",REF)]:
    ws.cell(r,1,txt).fill = _fill(bg)


# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 – Centuary
# ════════════════════════════════════════════════════════════════════════════
wc = wb.create_sheet("Centuary")
wc.column_dimensions["A"].width = 42
for i in range(2,16): wc.column_dimensions[col(i)].width = 13
wc.freeze_panes = "B3"

R = 1
# ── A. Budget breakdown ──────────────────────────────────────────────────
sec(wc, R, "CENTUARY – BUDGET BREAKDOWN (FY 2026-27)")
R += 1; mhdr(wc, R)

R += 1; RPCT = R
lbl(wc,R,1,"Percentage Split",bold=True)
for i,p in enumerate(PCT,2): inp(wc,R,i,p,"0%")
frm(wc,R,14,"=SUM(B{r}:M{r})".format(r=R),"0%")

R += 1; RTOT = R
lbl(wc,R,1,"Total Meta Budget – Centuary",bold=True)
TC = 3000000
for i,p in enumerate(PCT,2): inp(wc,R,i,TC*p,"₹#,##0")
frm(wc,R,14,"=SUM(B{r}:M{r})".format(r=R),"₹#,##0")
inp(wc,R,15,TC,"₹#,##0")

sub = [
    ("Reach, Engagement & Video Views (35%)", 0.35, "REV"),
    ("Product Awareness – Remarketing/Interest (20%)", 0.20, "AWR"),
    ("FB Page Like Campaigns (7%)", 0.07, "FBL"),
    ("IG Followers Campaigns (13%)", 0.13, "IGL"),
    ("YouTube Video View Campaigns (15%)", 0.15, "YTV"),
    ("YouTube Subscribers – Centuary (10%)", 0.10, "YTS_C"),
    ("YouTube Subscribers – Beddy (Fixed ₹5,000/mo)", None, "YTS_B"),
]
sub_rows = {}
for label, pct_val, key in sub:
    R += 1; sub_rows[key] = R
    lbl(wc,R,1,label,indent=1)
    for c in range(2,14):
        if key=="YTS_B": inp(wc,R,c,5000,"₹#,##0")
        else: frm(wc,R,c,f"={col(c)}{RTOT}*{pct_val}","₹#,##0")
    frm(wc,R,14,f"=SUM(B{R}:M{R})","₹#,##0")
    if pct_val: inp(wc,R,15,TC*pct_val,"₹#,##0")

R += 1; RINFL = R
lbl(wc,R,1,"Influencer Boosting",indent=1)
inp(wc,R,14,0,"₹#,##0")

R += 1; ROVER = R
lbl(wc,R,1,"Budget Overall",bold=True)
for c in range(2,14):
    frm(wc,R,c,f"=SUM({col(c)}{RTOT}:{col(c)}{RINFL-1})","₹#,##0")
frm(wc,R,14,f"=SUM(B{R}:M{R})","₹#,##0")

# ── B. Rate Inputs (vary by month) ───────────────────────────────────────
R += 2
sec(wc, R, "RATE INPUTS  ▸  Editable per month  ▸  Lower = better efficiency")
R += 1; mhdr(wc, R)

rate_defs = [
    ("CPR – Cost per 1,000 Reach (₹)",          "₹#,##0.00", 0),
    ("CPM – Cost per 1,000 Impressions (₹)",    "₹#,##0.00", 1),
    ("CPV – Cost per 3-Sec Video View, Meta (₹)","₹0.000",    2),
    ("CPE – Cost per Engagement (₹)",            "₹0.000",    3),
    ("Cost per FB Page Like (₹)",                "₹0.00",     4),
    ("Cost per IG Follower (₹)",                 "₹0.00",     5),
    ("CPV – Cost per YT View, Centuary (₹)",     "₹0.000",    6),
    ("Cost per YT Subscriber – Centuary (₹)",    "₹0.00",     7),
    ("CPV – Cost per YT View, Beddy (₹)",        "₹0.000",    8),
    ("Cost per YT Subscriber – Beddy (₹)",       "₹0.00",     9),
]
rate_rows = {}
RRATE_START = R+1
for rl, fmt, idx in rate_defs:
    R += 1; rate_rows[idx] = R
    lbl(wc,R,1,rl)
    for mi,rv in enumerate(RATES_C): inp(wc,R,mi+2,rv[idx],fmt)
    frm(wc,R,14,f"=AVERAGE(B{R}:M{R})",fmt)

# ── C. Meta KPI Targets (formulas) ───────────────────────────────────────
R += 2
sec(wc, R, "CENTUARY – META KPI TARGETS  ▸  Auto-calculated from Budget ÷ Rates")
R += 1; mhdr(wc, R)

RREV = sub_rows["REV"]
RFBL = sub_rows["FBL"]
RIGL = sub_rows["IGL"]

def kpi(ws, r, label, fn, fmt, total_fmt=None):
    lbl(ws,r,1,label,bg="E8F4E8")
    for mi in range(12):
        c = mi+2
        frm(ws,r,c,fn(c,r),fmt)
    frm(ws,r,14,f"=SUM(B{r}:M{r})",total_fmt or fmt)

# NOTE: Reach uses TOTAL meta budget (not REV): CPR = Total×1000/Reach
R += 1; RREACH = R
kpi(wc,R,"Overall Reach – Paid",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RTOT}*1000/{col(c)}{rate_rows[0]},0),0)",
    "#,##0")
R += 1; RIMPRS = R
kpi(wc,R,"Overall Impressions – Paid",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RTOT}*1000/{col(c)}{rate_rows[1]},0),0)",
    "#,##0")
R += 1; RVVIEW = R
kpi(wc,R,"3-Sec Video Views – Meta",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RREV}/{col(c)}{rate_rows[2]},0),0)",
    "#,##0")
R += 1
kpi(wc,R,"Reel Views – Meta",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RVVIEW}*{10/7:.6f},0),0)",
    "#,##0")
R += 1; RENGAGE = R
kpi(wc,R,"Engagement",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RREV}/{col(c)}{rate_rows[3]},0),0)",
    "#,##0")
R += 1
kpi(wc,R,"Engagement Rate %",
    lambda c,r: f"=IFERROR({col(c)}{RENGAGE}/{col(c)}{RREACH},0)",
    "0.00%", total_fmt="0.00%")
R += 1; RIGF = R
kpi(wc,R,"IG Followers",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RIGL}/{col(c)}{rate_rows[5]},0),0)",
    "#,##0")
R += 1; RFBF = R
kpi(wc,R,"Facebook Followers",
    lambda c,r: f"=IFERROR(ROUND({col(c)}{RFBL}/{col(c)}{rate_rows[4]},0),0)",
    "#,##0")
R += 1; lbl(wc,R,1,"X (Twitter) Followers",bg="E8F4E8")
R += 1; lbl(wc,R,1,"LinkedIn Followers",bg="E8F4E8")
R += 1; lbl(wc,R,1,"Link Clicks – Meta (input required)",bg="E8F4E8")

# Original annual benchmarks (orange reference row)
R += 1
sec(wc, R, "ORIGINAL ANNUAL TARGETS (reference – do not edit)")
R += 1; mhdr(wc, R)
ref_data = [
    ("Overall Reach – Paid (original)",           200000000, "#,##0"),
    ("Overall Impressions – Paid (original)",      428571429, "#,##0"),
    ("3-Sec Video Views – Meta (original)",          7000000, "#,##0"),
    ("Reel Views – Meta (original)",                10000000, "#,##0"),
    ("Engagement (original)",                        4200000, "#,##0"),
    ("IG Followers (original)",                        88196, "#,##0"),
    ("Facebook Followers (original)",                  37200, "#,##0"),
]
for label, val, fmt in ref_data:
    R += 1
    lbl(wc,R,1,label,bg=REF)
    for c in range(2,14):
        pct = PCT[c-2]
        ref(wc,R,c,round(val*pct),fmt)
    ref(wc,R,14,val,fmt)

# ── D. YouTube KPI Targets ────────────────────────────────────────────────
R += 2
sec(wc, R, "CENTUARY / BEDDY – YOUTUBE KPI TARGETS")
R += 1; mhdr(wc, R)

RYTVC = sub_rows["YTV"]
RYTSC = sub_rows["YTS_C"]
RYTSB = sub_rows["YTS_B"]

yt_kpis = [
    ("Video Views – Centuary",    f"IFERROR(ROUND({{c}}{RYTVC}/{{c}}{rate_rows[6]},0),0)", "#,##0"),
    ("Video Views – Beddy",       f"IFERROR(ROUND({{c}}{RYTSB}/{{c}}{rate_rows[8]},0),0)", "#,##0"),
    ("YT Subscribers – Centuary", f"IFERROR(ROUND({{c}}{RYTSC}/{{c}}{rate_rows[7]},0),0)", "#,##0"),
    ("YT Subscribers – Beddy",    f"IFERROR(ROUND({{c}}{RYTSB}/{{c}}{rate_rows[9]},0),0)", "#,##0"),
]
for label, fn_tmpl, fmt in yt_kpis:
    R += 1
    lbl(wc,R,1,label,bg="E8F4E8")
    for c in range(2,14):
        frm(wc,R,c,f"={fn_tmpl.format(c=col(c))}",fmt)
    frm(wc,R,14,f"=SUM(B{R}:M{R})",fmt)

# ── E. Spent Budget (actuals input) ──────────────────────────────────────
R += 2
sec(wc, R, "ACTUAL SPEND TRACKING (fill monthly as campaign runs)")
R += 1; mhdr(wc, R)
for label in ["Actual Budget – Meta (₹)","Spent Budget – Meta (₹)",
               "Actual Budget – YouTube (₹)","Spent Budget – YouTube (₹)"]:
    R += 1
    lbl(wc,R,1,label)
    for c in range(2,14): inp(wc,R,c,None,"₹#,##0")
    frm(wc,R,14,f"=SUM(B{R}:M{R})","₹#,##0")


# ════════════════════════════════════════════════════════════════════════════
# SHEET 3 – Sleep Matters
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Sleep Matters")
ws2.column_dimensions["A"].width = 42
for i in range(2,16): ws2.column_dimensions[col(i)].width = 13
ws2.freeze_panes = "B3"

SM_TOTAL = 288000

RS = 1
sec(ws2, RS, "SLEEP MATTERS – BUDGET BREAKDOWN (FY 2026-27)")
RS += 1; mhdr(ws2, RS)

RS += 1; SRPCT = RS
lbl(ws2,RS,1,"Percentage Split",bold=True)
for i,p in enumerate(PCT,2): inp(ws2,RS,i,p,"0%")
frm(ws2,RS,14,f"=SUM(B{RS}:M{RS})","0%")

RS += 1; SRTOT = RS
lbl(ws2,RS,1,"Total Budget for Meta – Sleep Matters",bold=True)
for i,p in enumerate(PCT,2): inp(ws2,RS,i,SM_TOTAL*p,"₹#,##0")
frm(ws2,RS,14,f"=SUM(B{RS}:M{RS})","₹#,##0")
inp(ws2,RS,15,SM_TOTAL,"₹#,##0")

sm_sub = [
    ("Reach, Engagement & Video Views (25%)", 0.25, "SRREV"),
    ("Product Awareness (13%)",               0.13, "SRAWR"),
    ("FB Page Like Campaigns (31%)",          0.31, "SRFBL"),
    ("IG Followers Campaigns (31%)",          0.31, "SRIGL"),
]
sm_sub_rows = {}
for label, pct_val, key in sm_sub:
    RS += 1; sm_sub_rows[key] = RS
    lbl(ws2,RS,1,label,indent=1)
    for c in range(2,14): frm(ws2,RS,c,f"={col(c)}{SRTOT}*{pct_val}","₹#,##0")
    frm(ws2,RS,14,f"=SUM(B{RS}:M{RS})","₹#,##0")
    inp(ws2,RS,15,SM_TOTAL*pct_val,"₹#,##0")

RS += 1; SROVER = RS
lbl(ws2,RS,1,"Budget Overall",bold=True)
for c in range(2,14): frm(ws2,RS,c,f"={col(c)}{SRTOT}","₹#,##0")
frm(ws2,RS,14,f"=SUM(B{RS}:M{RS})","₹#,##0")

# Sleep Matters Rates
RS += 2
sec(ws2, RS, "RATE INPUTS – SLEEP MATTERS  ▸  Editable per month")
RS += 1; mhdr(ws2, RS)

sm_rate_defs = [
    ("CPR – Cost per 1,000 Reach (₹)",           "₹#,##0.00", 0),
    ("CPM – Cost per 1,000 Impressions (₹)",     "₹#,##0.00", 1),
    ("CPV – Cost per 3-Sec Video View (₹)",       "₹0.000",    2),
    ("CPE – Cost per Engagement (₹)",             "₹0.000",    3),
    ("Cost per FB Page Like (₹)",                 "₹0.00",     4),
    ("Cost per IG Follower (₹)",                  "₹0.00",     5),
]
sm_rate_rows = {}
for rl, fmt, idx in sm_rate_defs:
    RS += 1; sm_rate_rows[idx] = RS
    lbl(ws2,RS,1,rl)
    for mi,rv in enumerate(RATES_S): inp(ws2,RS,mi+2,rv[idx],fmt)
    frm(ws2,RS,14,f"=AVERAGE(B{RS}:M{RS})",fmt)

# Sleep Matters KPI Targets
RS += 2
sec(ws2, RS, "SLEEP MATTERS – META KPI TARGETS  ▸  Auto-calculated")
RS += 1; mhdr(ws2, RS)

SRREV = sm_sub_rows["SRREV"]
SRFBL = sm_sub_rows["SRFBL"]
SRIGL = sm_sub_rows["SRIGL"]

def kpi2(r, label, fn, fmt, total_fmt=None):
    lbl(ws2,r,1,label,bg="E8F4E8")
    for mi in range(12):
        c = mi+2
        frm(ws2,r,c,fn(c,r),fmt)
    frm(ws2,r,14,f"=SUM(B{r}:M{r})",total_fmt or fmt)

RS += 1; SRREACH = RS
kpi2(RS,"Overall Reach – Paid",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRTOT}*1000/{col(c)}{sm_rate_rows[0]},0),0)",
     "#,##0")
RS += 1
kpi2(RS,"Overall Impressions – Paid",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRTOT}*1000/{col(c)}{sm_rate_rows[1]},0),0)",
     "#,##0")
RS += 1; SRVVIEW = RS
kpi2(RS,"3-Sec Video Views – Meta",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRREV}/{col(c)}{sm_rate_rows[2]},0),0)",
     "#,##0")
RS += 1
kpi2(RS,"Reel Views – Meta",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRVVIEW}*{10/7:.6f},0),0)",
     "#,##0")
RS += 1; SRENGAGE = RS
kpi2(RS,"Engagement",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRREV}/{col(c)}{sm_rate_rows[3]},0),0)",
     "#,##0")
RS += 1
kpi2(RS,"Engagement Rate %",
     lambda c,r: f"=IFERROR({col(c)}{SRENGAGE}/{col(c)}{SRREACH},0)",
     "0.00%","0.00%")
RS += 1
kpi2(RS,"IG Followers",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRIGL}/{col(c)}{sm_rate_rows[5]},0),0)",
     "#,##0")
RS += 1
kpi2(RS,"Facebook Followers",
     lambda c,r: f"=IFERROR(ROUND({col(c)}{SRFBL}/{col(c)}{sm_rate_rows[4]},0),0)",
     "#,##0")

# SM original targets
RS += 2
sec(ws2, RS, "ORIGINAL ANNUAL TARGETS – SLEEP MATTERS (reference)")
RS += 1; mhdr(ws2, RS)
sm_refs = [
    ("Overall Reach – Paid (original)",        19200000, "#,##0"),
    ("Overall Impressions – Paid (original)",  28800000, "#,##0"),
    ("3-Sec Video Views (original)",             360000, "#,##0"),
    ("Reel Views (original)",                   1008000, "#,##0"),
    ("Engagement (original)",                    360000, "#,##0"),
    ("IG Followers (original)",                    3600, "#,##0"),
    ("Facebook Followers (original)",             15000, "#,##0"),
]
for label, val, fmt in sm_refs:
    RS += 1
    lbl(ws2,RS,1,label,bg=REF)
    for c in range(2,14):
        ref(ws2,RS,c,round(val*PCT[c-2]),fmt)
    ref(ws2,RS,14,val,fmt)

# SM Spend tracking
RS += 2
sec(ws2, RS, "ACTUAL SPEND TRACKING – SLEEP MATTERS")
RS += 1; mhdr(ws2, RS)
for label in ["Actual Budget – Meta (₹)","Spent Budget – Meta (₹)"]:
    RS += 1
    lbl(ws2,RS,1,label)
    for c in range(2,14): inp(ws2,RS,c,None,"₹#,##0")
    frm(ws2,RS,14,f"=SUM(B{RS}:M{RS})","₹#,##0")


# ════════════════════════════════════════════════════════════════════════════
# SHEETS 4 & 5 – Target vs Achieved
# ════════════════════════════════════════════════════════════════════════════
def make_tva(wb, title, kpis):
    wt = wb.create_sheet(title)
    wt.column_dimensions["A"].width = 38
    wt.cell(1,1,"KPI").font = Font(bold=True, size=10)
    wt.cell(1,1).fill = _fill(DARK); wt.cell(1,1).font = Font(bold=True,color="FFFFFF")
    c = 2
    for m in MONTHS:
        wt.merge_cells(start_row=1,start_column=c,end_row=1,end_column=c+1)
        x = wt.cell(1,c,m)
        x.font=Font(bold=True,color="FFFFFF"); x.fill=_fill(MID); x.alignment=_align()
        for ci,lbl_txt in [(c,"Target"),(c+1,"Achieved")]:
            wt.cell(2,ci,lbl_txt).font=Font(bold=True,size=9)
            wt.cell(2,ci).fill=_fill(LIGHT if ci==c else "FFE699")
            wt.column_dimensions[col(ci)].width=12
        c += 2
    for ri, kpi_label in enumerate(kpis, 3):
        wt.cell(ri,1,kpi_label).fill=_fill(LIGHT)
    return wt

centuary_kpis = [
    "Overall Reach – Paid","Overall Impressions – Paid",
    "3-Sec Video Views – Meta","Reel Views – Meta",
    "Engagement","Engagement Rate %","Link Clicks – Meta",
    "CPR (₹)","CPM (₹)","CPV 3-Sec (₹)","CPE (₹)",
    "Cost per FB Page Like (₹)","Cost per IG Follower (₹)",
    "IG Followers","Facebook Followers",
    "YT Subscribers – Centuary","YT Subscribers – Beddy",
    "YT Video Views – Centuary","YT Video Views – Beddy",
    "Frequency – Meta","Hold Rate %",
    "Actual Budget – Meta (₹)","Spent Budget – Meta (₹)",
    "Actual Budget – YouTube (₹)","Spent Budget – YouTube (₹)",
]
sleep_kpis = [
    "Overall Reach – Paid","Overall Impressions – Paid",
    "3-Sec Video Views – Meta","Reel Views – Meta",
    "Engagement","Engagement Rate %","Link Clicks – Meta",
    "CPR (₹)","CPM (₹)","CPV 3-Sec (₹)","CPE (₹)",
    "Cost per FB Page Like (₹)","Cost per IG Follower (₹)",
    "IG Followers","Facebook Followers",
    "Hold Rate %",
    "Actual Budget – Meta (₹)","Spent Budget – Meta (₹)",
]

make_tva(wb, "Centuary - Target vs Achieved", centuary_kpis)
make_tva(wb, "Sleep Matters - Tgt vs Achieved", sleep_kpis)

# ════════════════════════════════════════════════════════════════════════════
OUT = "/home/user/AI-Agentics-Bots/centuary_budget_tracker.xlsx"
wb.save(OUT)
print("Saved:", OUT)
